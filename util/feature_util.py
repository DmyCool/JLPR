import pandas as pd
import numpy as np
from openpyxl import load_workbook
pd.set_option("display.max_columns", 1000)

# data processing
def generate_all_features():

    train_crime = np.load('../data/chicago-crime-tract-level-train.npy') #(801,)
    test_crime = np.load('../data/chicago-crime-tract-level-test.npy')
    train_crime, test_crime = list(train_crime), list(test_crime)

    price_train = pd.read_csv('../data/chicago-house-price-tract-level-train.csv', index_col=0)
    price_test = pd.read_csv('../data/chicago-house-price-tract-level-test.csv', index_col=0)

    train_price = list(price_train['train_price']) #(801,)
    train_count = list(price_train['train_count'])
    test_price = list(price_test['test_price'])
    test_count = list(price_test['test_count'])

    poi = list(np.load('../data/poi_emb1_mtx.npy'))  #(801, 32)
    img = list(np.load('../data/img_feat_mtx.npy')) #(801, 128)

    f, _ = retrieve_income_features()
    f_select = f['B1901001']
    population = f_select.sort_index().tolist()
    features = {'train_crime': train_crime,
                'test_crime': test_crime,
                'train_price': train_price,
                'train_count': train_count,
                'test_price': test_price,
                'test_count': test_count,
                'poi': poi,
                'img': img,
                'population': population}

    # covert to dataframe
    feats = pd.DataFrame(features)

    return feats

def retrieve_income_features():
    """
    read the xls file '../data/Household Income by Race and Census Tract and Community Area.xlsx'
    """
    wb = load_workbook('../data/Household Income by Race and Census Tract and Community Area.xlsx')
    ws = wb.active
    tractColumn = [cell.value for cell in ws['h']]
    dataColumns = ws['K1:DU890']

    header = []
    header_description = []
    income_features = []
    tractIDs = []
    for idx, tractID in enumerate(tractColumn):
        if idx == 0:
            header = [cell.value for cell in dataColumns[idx]]
        elif idx == 1:
            header_description = [cell.value for cell in dataColumns[idx]]
        elif idx == 2:
            header_description = ["{} {}".format(header_description[i], cell.value)
                                  for i, cell in enumerate(dataColumns[idx])]
        else:
            if tractID != None:
                tractIDs.append(int("17031"+tractID))
                row = [cell.value for cell in dataColumns[idx]]
                income_features.append(row)
    featureDF = pd.DataFrame(data=income_features, index=tractIDs, columns=header)
    featureDF = featureDF.loc[:,~featureDF.columns.duplicated()]
    header_decode = dict(zip(header, header_description))
    return featureDF, header_decode



if __name__ == '__main__':
    feats = generate_all_features()
    print(feats['img'][0])



